import io
import re
import openpyxl
import structlog
from uuid import uuid4
from datetime import datetime, timezone
from decimal import Decimal
from typing import Optional, List, Tuple

from app.domain.models import PastProject
from app.domain.repositories import (
    IPastProjectRepository,
    IProjectExtractor,
    IStorageProvider,
    IPDFExtractor,
    IEmbeddingProvider,
    IVectorSearchProvider,
)


logger = structlog.get_logger("app.projects")


class ProjectService:
    def __init__(
        self,
        project_repo: IPastProjectRepository,
        extractor: IProjectExtractor,
        storage: IStorageProvider,
        primary_extractor: IPDFExtractor,
        fallback_extractor: IPDFExtractor,
        embedding_provider: Optional[IEmbeddingProvider] = None,
        vector_search_provider: Optional[IVectorSearchProvider] = None,
    ):
        self.project_repo = project_repo
        self.extractor = extractor
        self.storage = storage
        self.primary_extractor = primary_extractor
        self.fallback_extractor = fallback_extractor
        self.embedding_provider = embedding_provider
        self.vector_search_provider = vector_search_provider


    async def extract_and_save_project(
        self, filename: str, content: bytes, doc_type: str
    ) -> PastProject:
        """Saves project document to local storage, extracts metadata, and stores record in database."""
        # Save file to disk
        saved_path = await self.storage.save(filename, content)
        logger.info("Saved project document", path=saved_path, filename=filename)

        raw_text = ""
        # 1. Primary Extractor (pdfplumber)
        try:
            logger.info("Attempting text extraction with primary extractor", path=saved_path)
            raw_text = await self.primary_extractor.extract_text(saved_path)
        except Exception as e:
            logger.warn("Primary text extraction failed, falling back", error=str(e), path=saved_path)

        # 2. Fallback Extractor (PyMuPDF)
        if not raw_text or not raw_text.strip():
            try:
                logger.info("Attempting text extraction with fallback extractor", path=saved_path)
                raw_text = await self.fallback_extractor.extract_text(saved_path)
            except Exception as e:
                logger.error("Fallback text extraction failed", error=str(e), path=saved_path)

        # 3. If extraction failed completely, use default empty text
        if not raw_text:
            raw_text = ""

        # 4. Extract structured details using rule-based provider
        extracted_details = await self.extractor.extract_project_details(raw_text, doc_type)

        # 5. Enforce numeric validations (negative values not allowed, reset to 0.00)
        project_value = extracted_details.get("project_value", Decimal("0.00"))
        if project_value < Decimal("0.0"):
            logger.warn(
                "Extracted negative project value, resetting to 0.00",
                value=str(project_value),
            )
            project_value = Decimal("0.00")

        # 6. Create Domain Entity
        now = datetime.now(timezone.utc)
        project = PastProject(
            id=uuid4(),
            project_name=extracted_details.get("project_name", "UNKNOWN"),
            client=extracted_details.get("client", "UNKNOWN"),
            project_value=project_value,
            completion_date=extracted_details.get("completion_date"),
            domain=extracted_details.get("domain", "Other"),
            location=extracted_details.get("location", "UNKNOWN"),
            document_type=doc_type,
            document_path=saved_path,
            created_at=now,
            updated_at=now,
        )

        # 7. Persist
        saved_project = await self.project_repo.add(project)
        logger.info(
            "Successfully extracted and saved past project credential",
            project_id=str(saved_project.id),
            project_name=saved_project.project_name,
        )

        # 8. Index in vector search if providers are registered
        if self.embedding_provider and self.vector_search_provider:
            try:
                await self.vector_search_provider.initialize()
                text = (
                    f"Project Name: {saved_project.project_name}. "
                    f"Client: {saved_project.client}. "
                    f"Domain: {saved_project.domain}. "
                    f"Location: {saved_project.location}."
                )
                logger.info("Generating embedding for automatic indexing", project_id=str(saved_project.id))
                embedding = await self.embedding_provider.embed_text(text, is_query=False)
                await self.vector_search_provider.upsert_project(saved_project, embedding)
                logger.info("Automatically indexed project in vector search", project_id=str(saved_project.id))
            except Exception as e:
                logger.error(
                    "Failed to automatically index project in vector search",
                    error=str(e),
                    project_id=str(saved_project.id),
                )

        return saved_project


    async def list_projects(
        self,
        skip: int = 0,
        limit: int = 10,
        search: Optional[str] = None,
        domain: Optional[str] = None,
        location: Optional[str] = None,
        min_value: Optional[Decimal] = None,
    ) -> Tuple[List[PastProject], int]:
        """Queries past projects with pagination, full-text search, and metadata filters."""
        return await self.project_repo.list(
            skip=skip,
            limit=limit,
            search=search,
            domain=domain,
            location=location,
            min_value=min_value,
        )

    async def get_capabilities(self) -> List[dict]:
        """Generates capability aggregates grouped by domain."""
        return await self.project_repo.get_capabilities()

    async def import_projects_from_excel(self, file_content: bytes) -> dict:
        """Imports past projects from an Excel sheet containing domain categories and metadata."""
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        sheet = wb.worksheets[0]
        
        active_domain = "Telecommunication"
        inserted = 0
        duplicates = 0
        failed = 0
        total_rows = 0
        errors = []
        
        for row_idx in range(2, sheet.max_row + 1):
            val_a = sheet.cell(row=row_idx, column=1).value
            val_b = sheet.cell(row=row_idx, column=2).value
            val_c = sheet.cell(row=row_idx, column=3).value
            val_d = sheet.cell(row=row_idx, column=4).value
            val_e = sheet.cell(row=row_idx, column=5).value
            val_f = sheet.cell(row=row_idx, column=6).value
            val_g = sheet.cell(row=row_idx, column=7).value
            val_h = sheet.cell(row=row_idx, column=8).value
            
            # Check if row is completely empty
            if all(v is None or str(v).strip() == "" for v in [val_a, val_b, val_c, val_d, val_e, val_f, val_g, val_h]):
                continue
                
            total_rows += 1
            
            # Heuristic for Domain category section row:
            is_section = False
            section_val = ""
            if val_b and not any(str(x).strip() for x in [val_a, val_c, val_e, val_f, val_g, val_h] if x is not None):
                is_section = True
                section_val = str(val_b).strip()
            elif val_d and not any(str(x).strip() for x in [val_a, val_b, val_c, val_e, val_f, val_g, val_h] if x is not None):
                is_section = True
                section_val = str(val_d).strip()
                
            if is_section:
                sec_lower = section_val.lower()
                if "signal" in sec_lower:
                    active_domain = "Signaling"
                elif "audio" in sec_lower or "video" in sec_lower or "video" in sec_lower:
                    active_domain = "Audio Video"
                elif "telecom" in sec_lower or "network" in sec_lower or "ofc" in sec_lower or "cable" in sec_lower:
                    active_domain = "Telecommunication"
                continue
                
            # We must have Project Name in Column D and Client/Location in Column B
            if not val_d or str(val_d).strip() == "":
                failed += 1
                errors.append({"row": row_idx, "error": "Missing Project Name in Column D"})
                continue
            if not val_b or str(val_b).strip() == "":
                failed += 1
                errors.append({"row": row_idx, "error": "Missing Client/Location in Column B"})
                continue
                
            project_name = str(val_d).strip()
            client = str(val_b).strip()
            location = str(val_b).strip()
            
            # Parse project value
            project_value = Decimal("0.00")
            if val_g is not None and str(val_g).strip() != "":
                try:
                    clean_val = re.sub(r'[^\d.]', '', str(val_g))
                    if clean_val:
                        project_value = Decimal(clean_val)
                except Exception as e:
                    failed += 1
                    errors.append({"row": row_idx, "project_name": project_name, "error": f"Invalid project value '{val_g}': {str(e)}"})
                    continue
                    
            # Parse completion date
            completion_date = None
            if val_h is not None and str(val_h).strip() != "":
                h_str = str(val_h).strip()
                if h_str.lower() not in ("running", "ongoing", "current", "-"):
                    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
                        try:
                            completion_date = datetime.strptime(h_str, fmt).date()
                            break
                        except Exception:
                            continue
                            
            doc_path = str(val_e).strip() if val_e is not None else None
            
            now = datetime.now(timezone.utc)
            project = PastProject(
                id=uuid4(),
                project_name=project_name,
                client=client,
                project_value=project_value,
                completion_date=completion_date,
                domain=active_domain,
                location=location,
                document_type="LOA",
                document_path=doc_path,
                created_at=now,
                updated_at=now
            )
            
            try:
                # Save to relational DB
                await self.project_repo.add(project)
                
                # Index in Qdrant Vector search
                if self.embedding_provider and self.vector_search_provider:
                    try:
                        await self.vector_search_provider.initialize()
                        text = (
                            f"Project Name: {project.project_name}. "
                            f"Client: {project.client}. "
                            f"Domain: {project.domain}. "
                            f"Location: {project.location}."
                        )
                        embedding = await self.embedding_provider.embed_text(text, is_query=False)
                        await self.vector_search_provider.upsert_project(project, embedding)
                    except Exception as ve:
                        logger.error("Failed to index project in vector search", project_id=str(project.id), error=str(ve))
                inserted += 1
            except Exception as db_err:
                failed += 1
                errors.append({"row": row_idx, "project_name": project_name, "error": f"Database save failed: {str(db_err)}"})
                
        return {
            "total_rows": total_rows,
            "inserted": inserted,
            "duplicates": duplicates,
            "failed": failed,
            "errors": errors
        }
